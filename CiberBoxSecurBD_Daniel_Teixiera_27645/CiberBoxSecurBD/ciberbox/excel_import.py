"""Leitura e validacao dos ficheiros Excel de ativos e incidentes.

A importacao aceita os dois modelos fornecidos com o enunciado e tambem os
modelos simplificados incluidos na pasta exemplos. A leitura nao grava dados:
primeiro e produzida uma pre-visualizacao e so depois da confirmacao as linhas
validas sao persistidas por SQL direto.
"""
from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path
from typing import Any
from uuid import uuid4
import ipaddress
import json
import re
import unicodedata

from django.conf import settings
from django.utils import timezone
from django.utils.text import get_valid_filename
from openpyxl import load_workbook


ASSET_ALIASES = {
    'numero_inventario': ['numero inventario', 'numero de inventario', 'n inventario', 'id'],
    'tipo_equipamento': ['tipo equipamento', 'tipo de equipamento', 'selecionar'],
    'nome': ['nome', 'nome do equipamento nome do software', 'nome equipamento', 'ativo'],
    'tipologia': ['tipologia'],
    'modelo_versao': ['modelo', 'versao', 'modelo versao'],
    'numero_serie': ['n serie', 'numero serie', 'numero de serie'],
    'fabricante': ['fabricante'],
    'localizacao': ['localizacao'],
    'sistema_operativo': ['sistema operativo'],
    'criticidade': ['criticidade', 'classificacao'],
    'endereco_ip': ['endereco ip', 'ip'],
    'endereco_mac': ['endereco hw', 'endereco mac', 'mac'],
    'fqdn': ['fqdn'],
    'servico_suportado': ['servico suportado'],
    'responsavel_nome': ['nome responsavel', 'responsavel nome'],
    'responsavel_contacto': ['contacto', 'contacto responsavel', 'responsavel contacto'],
    'unidade_organica': ['unidade organica', 'departamento'],
    'aplicacoes_servicos': ['aplicacoes servicos', 'aplicacoes', 'servicos'],
    'observacoes': ['observacoes'],
    'comunicado_cncs': ['ja foi comunicado ao cncs s n', 'comunicado cncs'],
    'programa_gestao_risco': ['programa de gestao de risco s n', 'programa gestao risco'],
}

INCIDENT_ALIASES = {
    'codigo': ['id', 'codigo'],
    'data_incidente': ['data do incidente', 'data incidente', 'data'],
    'hora_incidente': ['hora do incidente', 'hora incidente', 'hora'],
    'registado_por': ['registrado por', 'registado por'],
    'departamento': ['departamento'],
    'tipo_incidente': ['tipo incidente', 'tipo de incidente'],
    'descricao': ['descricao incidente', 'descricao do incidente', 'descricao'],
    'utilizadores_afetados': ['numero de utilizadores afetados', 'utilizadores afetados'],
    'dados_comprometidos': ['dados comprometidos'],
    'sistemas_afetados': ['sistemas afetados'],
    'origem_ataque': ['origem do ataque', 'origem ataque'],
    'ip_atacante': ['endereco ip do atacante', 'ip do atacante'],
    'analise_log': ['analise de log', 'analise log'],
    'resposta_imediata': ['resposta imediata'],
    'medidas_corretivas': ['medidas corretivas'],
    'entidades_internas': ['entidades internas'],
    'entidades_externas': ['entidades externas'],
    'gravidade': ['gravidade'],
    'probabilidade_reincidencia': ['probabilidade de reincidencia'],
    'recomendacoes': ['recomendacoes'],
    'encerrado_em': ['data de encerramento do incidente', 'data encerramento'],
    'responsavel_encerramento': ['responsavel pelo encerramento'],
    'estado': ['estado'],
}


def normalizar(texto: Any) -> str:
    if texto is None:
        return ''
    texto = str(texto).strip().lower().replace('\n', ' ')
    texto = ''.join(ch for ch in unicodedata.normalize('NFKD', texto) if not unicodedata.combining(ch))
    texto = re.sub(r'[^a-z0-9]+', ' ', texto)
    return re.sub(r'\s+', ' ', texto).strip()


def _mapear_cabecalhos(valores: list[Any], aliases: dict[str, list[str]]) -> dict[str, int]:
    normalizados = [normalizar(v) for v in valores]
    resultado: dict[str, int] = {}
    for campo, opcoes in aliases.items():
        opcoes_norm = {normalizar(opcao) for opcao in opcoes}
        for indice, cabecalho in enumerate(normalizados):
            if cabecalho in opcoes_norm:
                resultado[campo] = indice
                break
    return resultado


def _encontrar_cabecalho(ws, aliases: dict[str, list[str]], obrigatorios: set[str]) -> tuple[int, dict[str, int]]:
    melhor: tuple[int, dict[str, int]] | None = None
    for numero_linha, linha in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), start=1):
        mapa = _mapear_cabecalhos(list(linha), aliases)
        if obrigatorios.issubset(mapa):
            return numero_linha, mapa
        if melhor is None or len(mapa) > len(melhor[1]):
            melhor = (numero_linha, mapa)
    encontrados = ', '.join(sorted((melhor or (0, {}))[1])) or 'nenhum'
    raise ValueError(f'Nao foi encontrado um cabecalho valido. Campos reconhecidos: {encontrados}.')


def _valor(linha: tuple[Any, ...], mapa: dict[str, int], campo: str) -> Any:
    indice = mapa.get(campo)
    if indice is None or indice >= len(linha):
        return None
    valor = linha[indice]
    if isinstance(valor, str):
        valor = valor.strip()
    return valor if valor != '' else None


def _booleano(valor: Any) -> bool:
    if isinstance(valor, bool):
        return valor
    return normalizar(valor) in {'sim', 's', 'yes', 'y', 'true', '1', 'x'}


def _criticidade(valor: Any, default: str = 'MEDIA') -> str:
    mapa = {
        'residual': 'RESIDUAL', 'baixa': 'BAIXA', 'baixo': 'BAIXA',
        'media': 'MEDIA', 'medio': 'MEDIA', 'alta': 'ALTA', 'alto': 'ALTA',
        'critica': 'CRITICA', 'critico': 'CRITICA', 'c ativo critico': 'CRITICA',
        'nc ativo nao critico': 'BAIXA',
    }
    return mapa.get(normalizar(valor), default)


def _probabilidade(valor: Any) -> str | None:
    mapa = {'baixa': 'BAIXA', 'media': 'MEDIA', 'alta': 'ALTA'}
    return mapa.get(normalizar(valor))


def _validar_ip(valor: Any) -> str | None:
    if valor in (None, ''):
        return None
    texto = str(valor).strip()
    ipaddress.ip_address(texto)
    return texto


def _validar_mac(valor: Any) -> str | None:
    if valor in (None, ''):
        return None
    texto = str(valor).strip().replace('-', ':').lower()
    if not re.fullmatch(r'(?:[0-9a-f]{2}:){5}[0-9a-f]{2}', texto):
        raise ValueError('Endereco MAC invalido.')
    return texto


def _inteiro(valor: Any, default: int = 0) -> int:
    if valor in (None, ''):
        return default
    return int(float(valor))


def _data_hora(data_valor: Any, hora_valor: Any = None) -> datetime:
    if isinstance(data_valor, datetime):
        resultado = data_valor
        if hora_valor and resultado.time() == time(0, 0):
            if isinstance(hora_valor, time):
                resultado = datetime.combine(resultado.date(), hora_valor)
            elif isinstance(hora_valor, str):
                resultado = datetime.combine(resultado.date(), datetime.strptime(hora_valor.strip(), '%H:%M').time())
    elif isinstance(data_valor, date):
        hora = hora_valor if isinstance(hora_valor, time) else time(0, 0)
        resultado = datetime.combine(data_valor, hora)
    elif isinstance(data_valor, str):
        texto = data_valor.strip()
        resultado = None
        for formato in ('%Y-%m-%d %H:%M', '%Y-%m-%d', '%d/%m/%Y %H:%M', '%d/%m/%Y'):
            try:
                resultado = datetime.strptime(texto, formato)
                break
            except ValueError:
                continue
        if resultado is None:
            raise ValueError('Data invalida.')
        if hora_valor and resultado.time() == time(0, 0):
            if isinstance(hora_valor, time):
                resultado = datetime.combine(resultado.date(), hora_valor)
            elif isinstance(hora_valor, str):
                resultado = datetime.combine(resultado.date(), datetime.strptime(hora_valor.strip(), '%H:%M').time())
    else:
        raise ValueError('Data do incidente obrigatoria ou invalida.')
    if timezone.is_naive(resultado):
        resultado = timezone.make_aware(resultado, timezone.get_current_timezone())
    return resultado


def _sheet(workbook, tipo: str):
    preferidas = ['Lista de Ativos', 'ASSETS', 'Ativos'] if tipo == 'ATIVOS' else ['Incidents', 'Incidentes']
    por_nome_normalizado = {normalizar(nome): nome for nome in workbook.sheetnames}
    for nome in preferidas:
        real = por_nome_normalizado.get(normalizar(nome))
        if real:
            return workbook[real]
    return workbook[workbook.sheetnames[0]]


def analisar_excel(caminho: str | Path, tipo: str) -> dict[str, Any]:
    tipo = tipo.upper()
    if tipo not in {'ATIVOS', 'INCIDENTES'}:
        raise ValueError('Tipo de importacao invalido.')
    workbook = load_workbook(filename=str(caminho), read_only=True, data_only=True)
    ws = _sheet(workbook, tipo)
    try:
        if tipo == 'ATIVOS':
            linhas = _analisar_ativos(ws)
        else:
            linhas = _analisar_incidentes(ws)
    finally:
        workbook.close()
    return {
        'tipo': tipo,
        'folha': ws.title,
        'linhas': linhas,
        'validas': sum(1 for l in linhas if not l.get('erro')),
        'rejeitadas': sum(1 for l in linhas if l.get('erro')),
    }


def _analisar_ativos(ws) -> list[dict[str, Any]]:
    header_row, mapa = _encontrar_cabecalho(ws, ASSET_ALIASES, {'nome'})
    resultado = []
    for numero, linha in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(v not in (None, '') for v in linha):
            continue
        dados: dict[str, Any] = {}
        erro = None
        try:
            nome = _valor(linha, mapa, 'nome')
            if not nome:
                raise ValueError('O nome do ativo e obrigatorio.')
            dados = {
                'numero_inventario': str(_valor(linha, mapa, 'numero_inventario') or '').strip() or None,
                'tipo_equipamento': _valor(linha, mapa, 'tipo_equipamento'),
                'nome': str(nome).strip(),
                'tipologia': _valor(linha, mapa, 'tipologia'),
                'modelo_versao': _valor(linha, mapa, 'modelo_versao'),
                'numero_serie': _valor(linha, mapa, 'numero_serie'),
                'fabricante': _valor(linha, mapa, 'fabricante'),
                'localizacao': _valor(linha, mapa, 'localizacao'),
                'sistema_operativo': _valor(linha, mapa, 'sistema_operativo'),
                'criticidade': _criticidade(_valor(linha, mapa, 'criticidade')),
                'endereco_ip': _validar_ip(_valor(linha, mapa, 'endereco_ip')),
                'endereco_mac': _validar_mac(_valor(linha, mapa, 'endereco_mac')),
                'fqdn': _valor(linha, mapa, 'fqdn'),
                'servico_suportado': _valor(linha, mapa, 'servico_suportado'),
                'responsavel_nome': _valor(linha, mapa, 'responsavel_nome'),
                'responsavel_contacto': _valor(linha, mapa, 'responsavel_contacto'),
                'unidade_organica': _valor(linha, mapa, 'unidade_organica'),
                'aplicacoes_servicos': _valor(linha, mapa, 'aplicacoes_servicos'),
                'observacoes': _valor(linha, mapa, 'observacoes'),
                'comunicado_cncs': _booleano(_valor(linha, mapa, 'comunicado_cncs')),
                'programa_gestao_risco': _booleano(_valor(linha, mapa, 'programa_gestao_risco')),
            }
        except Exception as exc:
            erro = str(exc)
        resultado.append({'numero_linha': numero, 'dados': dados, 'erro': erro})
    if not resultado:
        raise ValueError('A folha nao contem linhas de dados.')
    return resultado


def _analisar_incidentes(ws) -> list[dict[str, Any]]:
    header_row, mapa = _encontrar_cabecalho(ws, INCIDENT_ALIASES, {'tipo_incidente', 'descricao'})
    resultado = []
    for numero, linha in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        preenchidos = [v for v in linha if v not in (None, '')]
        if not preenchidos:
            continue
        # Alguns modelos trazem apenas numeros de linha pre-preenchidos.
        # Esses placeholders nao representam incidentes e sao ignorados.
        if len(preenchidos) == 1 and _valor(linha, mapa, 'codigo') not in (None, ''):
            continue
        dados: dict[str, Any] = {}
        erro = None
        try:
            tipo_incidente = _valor(linha, mapa, 'tipo_incidente')
            descricao = _valor(linha, mapa, 'descricao')
            if not tipo_incidente or not descricao:
                raise ValueError('Tipo e descricao do incidente sao obrigatorios.')
            data_hora = _data_hora(_valor(linha, mapa, 'data_incidente'), _valor(linha, mapa, 'hora_incidente'))
            codigo_raw = _valor(linha, mapa, 'codigo')
            codigo = str(codigo_raw).strip() if codigo_raw not in (None, '') else f'IMP-{data_hora:%Y%m%d}-{numero:04d}'
            encerrado_raw = _valor(linha, mapa, 'encerrado_em')
            encerrado_em = _data_hora(encerrado_raw) if encerrado_raw else None
            estado = normalizar(_valor(linha, mapa, 'estado')).upper().replace(' ', '_')
            if not estado:
                estado = 'ENCERRADO' if encerrado_em else 'ABERTO'
            if estado not in {'ABERTO', 'EM_ANALISE', 'ENCERRADO'}:
                estado = 'EM_ANALISE'
            dados = {
                'codigo': codigo[:40],
                'data_hora_incidente': data_hora,
                'registado_por': _valor(linha, mapa, 'registado_por'),
                'departamento': _valor(linha, mapa, 'departamento'),
                'tipo_incidente': str(tipo_incidente)[:100],
                'descricao': str(descricao),
                'utilizadores_afetados': _inteiro(_valor(linha, mapa, 'utilizadores_afetados')),
                'dados_comprometidos': _booleano(_valor(linha, mapa, 'dados_comprometidos')),
                'sistemas_afetados': _valor(linha, mapa, 'sistemas_afetados'),
                'origem_ataque': _valor(linha, mapa, 'origem_ataque'),
                'ip_atacante': _validar_ip(_valor(linha, mapa, 'ip_atacante')),
                'analise_log': _valor(linha, mapa, 'analise_log'),
                'resposta_imediata': _valor(linha, mapa, 'resposta_imediata'),
                'medidas_corretivas': _valor(linha, mapa, 'medidas_corretivas'),
                'entidades_internas': _valor(linha, mapa, 'entidades_internas'),
                'entidades_externas': _valor(linha, mapa, 'entidades_externas'),
                'gravidade': _criticidade(_valor(linha, mapa, 'gravidade')),
                'probabilidade_reincidencia': _probabilidade(_valor(linha, mapa, 'probabilidade_reincidencia')),
                'recomendacoes': _valor(linha, mapa, 'recomendacoes'),
                'estado': estado,
                'encerrado_em': encerrado_em,
                'responsavel_encerramento': _valor(linha, mapa, 'responsavel_encerramento'),
            }
        except Exception as exc:
            erro = str(exc)
        resultado.append({'numero_linha': numero, 'dados': dados, 'erro': erro})
    if not resultado:
        raise ValueError('A folha nao contem linhas de dados.')
    return resultado


def guardar_upload_excel(ficheiro) -> tuple[str, str]:
    pasta = Path(settings.MEDIA_ROOT) / 'private_uploads' / 'importacoes'
    pasta.mkdir(parents=True, exist_ok=True)
    nome_original = get_valid_filename(Path(ficheiro.name).name)
    nome_guardado = f'{uuid4().hex}.xlsx'
    caminho = pasta / nome_guardado
    with caminho.open('wb') as destino:
        for bloco in ficheiro.chunks():
            destino.write(bloco)
    relativo = caminho.relative_to(settings.MEDIA_ROOT).as_posix()
    return nome_original, relativo


def guardar_preview(conteudo: dict[str, Any]) -> str:
    pasta = Path(settings.MEDIA_ROOT) / 'import_previews'
    pasta.mkdir(parents=True, exist_ok=True)
    token = uuid4().hex
    caminho = pasta / f'{token}.json'
    caminho.write_text(json.dumps(conteudo, ensure_ascii=False, default=str), encoding='utf-8')
    return token


def carregar_preview(token: str) -> dict[str, Any]:
    if not re.fullmatch(r'[0-9a-f]{32}', token or ''):
        raise ValueError('Token de pre-visualizacao invalido.')
    caminho = Path(settings.MEDIA_ROOT) / 'import_previews' / f'{token}.json'
    return json.loads(caminho.read_text(encoding='utf-8'))


def apagar_preview(token: str) -> None:
    caminho = Path(settings.MEDIA_ROOT) / 'import_previews' / f'{token}.json'
    caminho.unlink(missing_ok=True)
